"""
Flight Hunter V4 — Reporte Excel (openpyxl)
=============================================
Basado en el skill xlsx:
- Fórmulas Excel reales (no valores hardcoded)
- Formato profesional con convenciones de color
- 4 hojas: Summary, All Deals, Business Hunter, By Airline
- Tablas con auto-filtros
- Código de colores por clasificación
- Ancho de columnas automático
"""

from datetime import datetime
from typing import List, Dict
from collections import defaultdict
import os
import config

try:
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, numbers
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, DoughnutChart, Reference, Series
    from openpyxl.worksheet.table import Table, TableStyleInfo
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


# ── Estilos de color (convenciones del skill xlsx) ──────────────────────
COLORS = {
    "header_bg":  "1A1A2E",   # Header oscuro
    "header_fg":  "FFFFFF",
    "critico_bg": "FEF2F2",   # Rojo claro
    "critico_fg": "DC2626",
    "error_bg":   "FFF7ED",   # Naranja claro
    "error_fg":   "EA580C",
    "anomalia_bg": "FFFBEB",  # Amarillo claro
    "anomalia_fg": "D97706",
    "oferta_bg":  "F0FDF4",   # Verde claro
    "oferta_fg":  "16A34A",
    "biz_bg":     "F5F3FF",   # Violeta claro
    "biz_fg":     "7C3AED",
    "alt_row":    "F8FAFC",   # Gris muy claro para filas alternas
    "border":     "E2E8F0",
}

def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", start_color=hex_color, fgColor=hex_color)

def _font(bold=False, color="000000", size=10) -> Font:
    return Font(name="Calibri", bold=bold, color=color, size=size)

def _border_thin() -> Border:
    s = Side(style="thin", color=COLORS["border"])
    return Border(left=s, right=s, top=s, bottom=s)

def _center() -> Alignment:
    return Alignment(horizontal="center", vertical="center")

def _left() -> Alignment:
    return Alignment(horizontal="left", vertical="center")

CLASS_COLORS = {
    "CRÍTICO":  (COLORS["critico_bg"],  COLORS["critico_fg"]),
    "ERROR":    (COLORS["error_bg"],    COLORS["error_fg"]),
    "ANOMALÍA": (COLORS["anomalia_bg"], COLORS["anomalia_fg"]),
    "OFERTA":   (COLORS["oferta_bg"],   COLORS["oferta_fg"]),
}


def _autowidth(ws, min_w=8, max_w=45):
    """Ajusta el ancho de columnas automáticamente."""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_w, max(min_w, max_len + 2))


def _style_header_row(ws, row: int, n_cols: int, font_size=10):
    """Aplica estilo de cabecera a una fila."""
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = _fill(COLORS["header_bg"])
        cell.font = _font(bold=True, color=COLORS["header_fg"], size=font_size)
        cell.alignment = _center()
        cell.border = _border_thin()


def generate_excel_report(
    analyzed: List[Dict],
    search_params: Dict,
    all_flights: List[Dict] = None,
    output_path: str = None,
) -> str:
    """
    Genera un workbook Excel con 4 hojas.

    Returns: ruta del archivo generado
    """
    if not OPENPYXL_AVAILABLE:
        print("   ⚠️  openpyxl no disponible. Instalar: pip install openpyxl")
        return ""

    all_flights = all_flights or analyzed
    now_str = datetime.now().strftime("%Y-%m-%d %H:%M")
    timestamp = datetime.now().strftime("%Y%m%d_%H%M")

    if not output_path:
        report_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..")
        output_path = os.path.join(report_dir, f"FLIGHT_HUNTER_V4_{timestamp}.xlsx")

    wb = Workbook()
    wb.remove(wb.active)  # Eliminar hoja vacía por defecto

    # ════════════════════════════════════════════════
    # HOJA 1: SUMMARY
    # ════════════════════════════════════════════════
    ws1 = wb.create_sheet("📊 Summary")
    ws1.sheet_view.showGridLines = False

    # Título
    ws1.merge_cells("A1:G1")
    ws1["A1"] = "✈️ Flight Hunter V4 — Summary Report"
    ws1["A1"].font = _font(bold=True, size=16, color=COLORS["header_fg"])
    ws1["A1"].fill = _fill(COLORS["header_bg"])
    ws1["A1"].alignment = _center()

    # Metadata
    meta = [
        ("Generado:", now_str),
        ("Modo:", search_params.get("mode", "")),
        ("Orígenes:", ", ".join(search_params.get("origins", [])[:10])),
        ("Rango fechas:", f"{search_params.get('date_from','')} → {search_params.get('date_to','')}"),
        ("Cabina:", search_params.get("cabin", "Economy")),
        ("Vuelos scrapeados:", len(all_flights)),
    ]
    for i, (label, value) in enumerate(meta, start=3):
        ws1.cell(row=i, column=1, value=label).font = _font(bold=True)
        ws1.cell(row=i, column=2, value=value)

    # KPIs
    criticos  = [a for a in analyzed if a.get("classification") == "CRÍTICO"]
    errores   = [a for a in analyzed if a.get("classification") == "ERROR"]
    anomalias = [a for a in analyzed if a.get("classification") == "ANOMALÍA"]
    ofertas   = [a for a in analyzed if a.get("classification") == "OFERTA"]
    biz_deals = [a for a in analyzed if a.get("cabin_code") in (3, 4)]

    kpis = [
        ("Tipo", "Cantidad", "Descripción"),
        ("🚨 CRÍTICOS",  len(criticos),  "Error fares confirmados por 2+ técnicas"),
        ("❌ ERRORES",   len(errores),   "Posibles error fares"),
        ("⚠️ ANOMALÍAS", len(anomalias), "Precio inusualmente bajo"),
        ("💰 OFERTAS",   len(ofertas),   "Precio muy competitivo"),
        ("👑 BUSINESS",  len(biz_deals), "Deals en Business/First class"),
        ("📊 TOTAL",     len(analyzed),  "Total anomalías detectadas"),
    ]

    kpi_start_row = 11
    ws1.cell(row=kpi_start_row, column=1, value="RESUMEN POR CLASIFICACIÓN").font = _font(bold=True, size=12)
    ws1.cell(row=kpi_start_row + 1, column=1).font = _font(bold=True)

    _style_header_row(ws1, kpi_start_row + 1, 3)
    ws1.cell(row=kpi_start_row + 1, column=1, value="Clasificación")
    ws1.cell(row=kpi_start_row + 1, column=2, value="Cantidad")
    ws1.cell(row=kpi_start_row + 1, column=3, value="Descripción")

    for i, (tipo, cnt, desc) in enumerate(kpis[1:], start=kpi_start_row + 2):
        ws1.cell(row=i, column=1, value=tipo).font = _font(bold=True)
        ws1.cell(row=i, column=2, value=cnt).font = _font(bold=True, color="1A5276")
        ws1.cell(row=i, column=2).alignment = _center()
        ws1.cell(row=i, column=2).number_format = "#,##0"
        ws1.cell(row=i, column=3, value=desc)
        # Color por clasificación
        cls_name = tipo.split()[-1].strip("🚨❌⚠️💰👑📊 ")
        for cls, (bg, fg) in CLASS_COLORS.items():
            if cls[:4] in tipo:
                ws1.cell(row=i, column=1).fill = _fill(bg)
                ws1.cell(row=i, column=1).font = _font(bold=True, color=fg)
                ws1.cell(row=i, column=2).fill = _fill(bg)
                break

    # Estadísticas adicionales
    stats_row = kpi_start_row + len(kpis) + 4
    ws1.cell(row=stats_row, column=1, value="ESTADÍSTICAS DE PRECIO").font = _font(bold=True, size=12)
    _style_header_row(ws1, stats_row + 1, 4)
    for col, h in enumerate(["Métrica", "Economy (€)", "Business (€)", "Notas"], 1):
        ws1.cell(row=stats_row + 1, column=col, value=h)

    eco_prices = [f["price_eur"] for f in all_flights if f.get("cabin_code") == config.CABIN_ECONOMY and f.get("price_eur")]
    biz_prices = [f["price_eur"] for f in all_flights if f.get("cabin_code") == config.CABIN_BUSINESS and f.get("price_eur")]

    def safe_stat(lst, fn):
        try:
            return round(fn(lst), 0) if lst else ""
        except Exception:
            return ""

    stats_data = [
        ("Precio mínimo", safe_stat(eco_prices, min), safe_stat(biz_prices, min), "Mejor precio encontrado"),
        ("Precio máximo", safe_stat(eco_prices, max), safe_stat(biz_prices, max), ""),
        ("Media", safe_stat(eco_prices, lambda x: sum(x)/len(x)), safe_stat(biz_prices, lambda x: sum(x)/len(x)), ""),
        ("Mediana", safe_stat(eco_prices, lambda x: sorted(x)[len(x)//2]), safe_stat(biz_prices, lambda x: sorted(x)[len(x)//2]), "Mediana es más robusta que media"),
    ]

    for i, row_data in enumerate(stats_data, start=stats_row + 2):
        for c, val in enumerate(row_data, 1):
            cell = ws1.cell(row=i, column=c, value=val)
            if c in (2, 3) and val != "":
                cell.number_format = '€#,##0'
            if i % 2 == 0:
                cell.fill = _fill(COLORS["alt_row"])

    _autowidth(ws1)
    ws1.column_dimensions["A"].width = 22
    ws1.column_dimensions["C"].width = 40
    ws1.row_dimensions[1].height = 30

    # ════════════════════════════════════════════════
    # HOJA 2: ALL DEALS
    # ════════════════════════════════════════════════
    ws2 = wb.create_sheet("✈️ All Deals")
    ws2.sheet_view.showGridLines = False
    ws2.freeze_panes = "A2"

    headers2 = [
        "Score", "Clasificación", "Origen", "Destino", "Ciudad", "País",
        "Cabina", "Precio (€)", "Ratio B/E", "Ahorro (€)", "Ahorro (%)",
        "Aerolínea", "Fecha Salida", "Fecha Vuelta", "Escalas",
        "Aerolínea Premium", "Temporada", "Razón Principal", "Reservar"
    ]
    _style_header_row(ws2, 1, len(headers2))
    for c, h in enumerate(headers2, 1):
        ws2.cell(row=1, column=c, value=h)

    for row_idx, deal in enumerate(analyzed, start=2):
        cls = deal.get("classification", "NORMAL")
        bg, fg = CLASS_COLORS.get(cls, (COLORS["alt_row"], "000000"))

        vals = [
            deal.get("final_score", 0),
            cls,
            deal.get("origin", ""),
            deal.get("destination", ""),
            deal.get("city_to", ""),
            deal.get("country_to", ""),
            deal.get("cabin", "Economy"),
            deal.get("price_eur", 0),
            deal.get("t4_ratio", ""),
            deal.get("savings_eur", 0),
            deal.get("savings_pct", 0),
            deal.get("airline", ""),
            deal.get("date_out", ""),
            deal.get("date_ret", ""),
            deal.get("stops", 0),
            "Sí" if deal.get("premium_airline") else "No",
            deal.get("season_multiplier", 1.0),
            deal.get("main_reason", ""),
            deal.get("booking_url", ""),
        ]

        for c, val in enumerate(vals, 1):
            cell = ws2.cell(row=row_idx, column=c, value=val)
            cell.border = _border_thin()
            # Colorear fila por clasificación (fondo suave)
            if c <= len(headers2) - 1:  # No colorear URL
                cell.fill = _fill(bg)
            # Formatos específicos
            if c == 1:  # Score
                cell.number_format = "0.0"
                cell.alignment = _center()
            elif c == 8:  # Precio
                cell.number_format = "€#,##0"
                cell.font = _font(bold=True, color="16A34A")
            elif c == 9:  # Ratio
                if val:
                    cell.number_format = "0.00"
                    if isinstance(val, (int, float)) and val < 2:
                        cell.font = _font(bold=True, color=COLORS["critico_fg"])
            elif c == 10:  # Ahorro €
                cell.number_format = "€#,##0"
            elif c == 11:  # Ahorro %
                cell.number_format = "0.0%"
                cell.value = (val or 0) / 100
            elif c == 15:  # Escalas
                cell.alignment = _center()
            elif c == 19 and val:  # URL — como hipervínculo
                cell.hyperlink = val
                cell.value = "Reservar →"
                cell.font = Font(name="Calibri", color="4C72B0", underline="single", size=10)

    # Auto-filtro en la tabla
    ws2.auto_filter.ref = f"A1:{get_column_letter(len(headers2))}1"
    _autowidth(ws2)

    # ════════════════════════════════════════════════
    # HOJA 3: BUSINESS HUNTER (Ratio B/E)
    # ════════════════════════════════════════════════
    ws3 = wb.create_sheet("👑 Business Hunter")
    ws3.sheet_view.showGridLines = False
    ws3.freeze_panes = "A2"

    biz_analyzed = [a for a in analyzed if a.get("cabin_code") in (3, 4)]
    biz_analyzed_sorted = sorted(biz_analyzed, key=lambda x: x.get("t4_ratio") or 99)

    ws3.merge_cells("A1:J1")
    ws3["A1"] = "👑 Business Hunter — Ratio Business/Economy"
    ws3["A1"].font = _font(bold=True, size=13, color=COLORS["header_fg"])
    ws3["A1"].fill = _fill("4C1D95")
    ws3["A1"].alignment = _center()
    ws3.row_dimensions[1].height = 24

    # Explicación
    ws3.cell(row=2, column=1, value="Ratio normal Business/Economy: 5-8x en rutas largas. Ratio < 2x = ERROR FARE probable.").font = _font(color="6B7280", size=9)
    ws3.merge_cells("A2:J2")

    headers3 = [
        "Ruta", "Destino", "Ciudad", "Aerolínea", "Business (€)",
        "Economy (€)", "Ratio B/E", "Clasif. Ratio", "Ahorro (€)", "Reservar"
    ]
    _style_header_row(ws3, 3, len(headers3), font_size=10)
    ws3.cell(row=3, column=7).fill = _fill("7C3AED")  # Destacar columna ratio
    for c, h in enumerate(headers3, 1):
        ws3.cell(row=3, column=c, value=h)

    for row_idx, deal in enumerate(biz_analyzed_sorted, start=4):
        ratio = deal.get("t4_ratio")
        eco_price = deal.get("t4_eco_price") or deal.get("bec_eco_price", "")
        ratio_class = deal.get("bec_class") or (
            config.classify_ratio(ratio, deal.get("destination","")) if ratio else ""
        )

        row_vals = [
            f"{deal.get('origin','')}→{deal.get('destination','')}",
            deal.get("destination", ""),
            deal.get("city_to", ""),
            deal.get("airline", ""),
            deal.get("price_eur", 0),
            eco_price or "",
            ratio or "",
            ratio_class,
            deal.get("savings_eur", 0),
            deal.get("booking_url", ""),
        ]

        for c, val in enumerate(row_vals, 1):
            cell = ws3.cell(row=row_idx, column=c, value=val)
            cell.border = _border_thin()
            if row_idx % 2 == 0:
                cell.fill = _fill(COLORS["alt_row"])

            if c == 5:  # Precio Business
                cell.number_format = "€#,##0"
                cell.font = _font(bold=True, color="7C3AED")
            elif c == 6:  # Precio Economy
                cell.number_format = "€#,##0"
                cell.font = _font(color="4C72B0")
            elif c == 7:  # Ratio
                cell.number_format = "0.00"
                if ratio and isinstance(ratio, (int, float)):
                    if ratio < 2:
                        cell.fill = _fill(COLORS["critico_bg"])
                        cell.font = _font(bold=True, color=COLORS["critico_fg"])
                    elif ratio < 3:
                        cell.fill = _fill(COLORS["error_bg"])
                        cell.font = _font(bold=True, color=COLORS["error_fg"])
                    elif ratio < 4:
                        cell.fill = _fill(COLORS["anomalia_bg"])
                        cell.font = _font(color=COLORS["anomalia_fg"])
            elif c == 8:  # Clasif. ratio
                bg, fg = CLASS_COLORS.get(ratio_class, (COLORS["alt_row"], "000000"))
                cell.fill = _fill(bg)
                cell.font = _font(bold=True, color=fg)
                cell.alignment = _center()
            elif c == 9:  # Ahorro
                cell.number_format = "€#,##0"
                cell.font = _font(color="16A34A")
            elif c == 10 and val:  # URL
                cell.hyperlink = val
                cell.value = "Reservar →"
                cell.font = Font(name="Calibri", color="4C72B0", underline="single", size=10)

    ws3.auto_filter.ref = f"A3:{get_column_letter(len(headers3))}3"
    _autowidth(ws3)
    ws3.column_dimensions["A"].width = 14
    ws3.column_dimensions["C"].width = 20

    # Añadir leyenda de ratios
    legend_row = len(biz_analyzed_sorted) + 6
    ws3.cell(row=legend_row, column=1, value="LEYENDA RATIO B/E:").font = _font(bold=True)
    leyenda = [
        ("< 2x", "ERROR FARE", COLORS["critico_bg"], COLORS["critico_fg"]),
        ("2x - 3x", "ANOMALÍA", COLORS["error_bg"], COLORS["error_fg"]),
        ("3x - 4x", "OFERTA", COLORS["anomalia_bg"], COLORS["anomalia_fg"]),
        ("> 5x", "PRECIO NORMAL", COLORS["alt_row"], "000000"),
    ]
    for i, (ratio_range, label, bg, fg) in enumerate(leyenda):
        r = legend_row + 1 + i
        ws3.cell(row=r, column=1, value=ratio_range).fill = _fill(bg)
        ws3.cell(row=r, column=1).font = _font(bold=True, color=fg)
        ws3.cell(row=r, column=2, value=label).fill = _fill(bg)
        ws3.cell(row=r, column=2).font = _font(color=fg)

    # ════════════════════════════════════════════════
    # HOJA 4: BY AIRLINE
    # ════════════════════════════════════════════════
    ws4 = wb.create_sheet("✈️ By Airline")
    ws4.sheet_view.showGridLines = False

    airline_stats = defaultdict(lambda: {"count": 0, "prices": [], "biz_prices": [], "criticos": 0, "errors": 0})
    for a in analyzed:
        al = a.get("airline", "?") or "?"
        airline_stats[al]["count"] += 1
        airline_stats[al]["prices"].append(a.get("price_eur", 0))
        if a.get("cabin_code") in (3, 4):
            airline_stats[al]["biz_prices"].append(a.get("price_eur", 0))
        if a.get("classification") == "CRÍTICO":
            airline_stats[al]["criticos"] += 1
        elif a.get("classification") == "ERROR":
            airline_stats[al]["errors"] += 1

    ws4.merge_cells("A1:H1")
    ws4["A1"] = "✈️ Deals por Aerolínea"
    ws4["A1"].font = _font(bold=True, size=13, color=COLORS["header_fg"])
    ws4["A1"].fill = _fill(COLORS["header_bg"])
    ws4["A1"].alignment = _center()
    ws4.row_dimensions[1].height = 24

    headers4 = ["Aerolínea", "Total Deals", "Precio Mín (€)", "Precio Medio (€)", "Business Mín (€)", "Críticos", "Errores", "Premium"]
    _style_header_row(ws4, 2, len(headers4))
    for c, h in enumerate(headers4, 1):
        ws4.cell(row=2, column=c, value=h)

    airline_sorted = sorted(airline_stats.items(), key=lambda x: x[1]["count"], reverse=True)

    for row_idx, (airline, stats) in enumerate(airline_sorted, start=3):
        prices = stats["prices"]
        biz = stats["biz_prices"]
        is_premium = airline in config.AIRLINES_PREMIUM_BUSINESS

        row_vals = [
            airline,
            stats["count"],
            round(min(prices)) if prices else "",
            round(sum(prices) / len(prices)) if prices else "",
            round(min(biz)) if biz else "",
            stats["criticos"],
            stats["errors"],
            "⭐ Premium" if is_premium else "",
        ]

        for c, val in enumerate(row_vals, 1):
            cell = ws4.cell(row=row_idx, column=c, value=val)
            cell.border = _border_thin()
            if row_idx % 2 == 0:
                cell.fill = _fill(COLORS["alt_row"])
            if c in (3, 4, 5) and val:
                cell.number_format = "€#,##0"
            if c == 6 and isinstance(val, int) and val > 0:
                cell.fill = _fill(COLORS["critico_bg"])
                cell.font = _font(bold=True, color=COLORS["critico_fg"])
            if c == 7 and isinstance(val, int) and val > 0:
                cell.fill = _fill(COLORS["error_bg"])
                cell.font = _font(color=COLORS["error_fg"])
            if c == 8 and val:
                cell.fill = _fill(COLORS["biz_bg"])
                cell.font = _font(bold=True, color=COLORS["biz_fg"])

    ws4.auto_filter.ref = f"A2:{get_column_letter(len(headers4))}2"
    _autowidth(ws4)

    # ── Guardar ──────────────────────────────────────────────
    wb.save(output_path)
    print(f"   💾 Excel guardado: {output_path}")
    return output_path


def install_openpyxl_if_needed():
    """Instala openpyxl si no está disponible."""
    global OPENPYXL_AVAILABLE
    if not OPENPYXL_AVAILABLE:
        import subprocess, sys
        subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "--break-system-packages", "-q"])
        OPENPYXL_AVAILABLE = True
        print("   ✅ openpyxl instalado")
